"""Excel adapter - turn a customer requirement spreadsheet into canonical form.

In real programmes requirements arrive as .xlsx at least as often as ReqIF, and
every supplier names the columns differently. So row extraction is
deterministic and only the column SEMANTICS are inferred by the model.

The rule that must not be broken: if the sheet carries the customer's own
requirement identifiers, they are PRESERVED. Renaming CUST-SYS-4711 to REQ-001
destroys upward traceability, which is the entire point of the matrix. Any
identifier we invent is flagged as derived, because a derived requirement
carries less weight in an audit than one that quotes the customer's number.
"""
import io, os, json, re
from openpyxl import load_workbook
from google import genai
from common import llm

PROJECT = os.environ["GCP_PROJECT"]
MODEL = os.environ.get("GEMINI_MODEL", "gemini-3.5-flash")
_client = genai.Client(vertexai=True, project=PROJECT, location="global")

CANONICAL = ["id", "text", "type", "status", "rationale", "parent"]
MAX_ROWS = 5000


def _cells(row):
    return ["" if c is None else str(c).strip() for c in row]


def find_header_row(rows, max_scan=25):
    """Real files carry titles, logos and merged cells above the actual table.

    Deterministic heuristic: a header row has several short labels and is
    followed by a row that actually has data.
    """
    best, best_score = None, 0
    for i, row in enumerate(rows[:max_scan]):
        filled = [c for c in row if c]
        if len(filled) < 2:
            continue
        shortish = sum(1 for c in filled if 0 < len(c) <= 40)
        below = rows[i + 1] if i + 1 < len(rows) else []
        data_below = sum(1 for c in below if c) >= 2
        score = shortish + (4 if data_below else 0)
        if score > best_score:
            best, best_score = i, score
    return best


def looks_like_heading(text, row, cols):
    """Section headings are rows of text with nothing else on them.

    A real requirement row almost always carries a type, a status or an id.
    A heading carries only a short label, often numbered. Deterministic, and
    cheaper and safer than asking a model row by row.
    """
    if re.match(r"^\d+(\.\d+)*[\s.)-]", text) and len(text) < 80:
        return True
    others = [f for f in ("id", "type", "status", "rationale", "parent")
              if cols.get(f) is not None]
    filled = [f for f in others
              if cols[f] < len(row) and row[cols[f]].strip()]
    return not filled and len(text) < 80


def map_columns(headers, samples):
    """Ask the model what each column MEANS. Parsing cannot answer that.

    One supplier writes 'ID', another 'Ident.', another 'Requirement Number',
    another it in German. Semantics is exactly the job a model should do, and
    exactly the job a regular expression should not.
    """
    prompt = (
        "You are reading a customer requirements spreadsheet. Map each column "
        "to a canonical field, using the header labels and the sample rows. "
        "Fields: id (the customer's own requirement identifier), text (the "
        "requirement wording), type, status, rationale, parent (identifier of "
        "a parent requirement). Use null when no column fits a field; never "
        "force a match. Column indexes are zero-based. "
        'Answer ONLY with JSON: {"id":0,"text":2,"type":null,"status":null,'
        '"rationale":null,"parent":null}. Always answer in English.\n\n'
        f"HEADERS: {json.dumps(headers)}\n\nSAMPLE ROWS:\n"
        + json.dumps(samples[:4]))
    raw = llm.generate(prompt).strip()
    for fence in ("```json", "```"):
        if raw.startswith(fence):
            raw = raw[len(fence):]
    if raw.endswith("```"):
        raw = raw[:-3]
    mapping = json.loads(raw.strip())
    return {k: mapping.get(k) for k in CANONICAL}


def parse(data, sheet_name=None):
    """Entry point. `data` is the raw bytes of an .xlsx file."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = [_cells(r) for r in sheet.iter_rows(max_row=MAX_ROWS, values_only=True)]
    if not rows:
        raise ValueError("the sheet is empty")

    h = find_header_row(rows)
    if h is None:
        raise ValueError("no header row could be identified")
    headers = rows[h]
    body = [r for r in rows[h + 1:] if any(c for c in r)]
    if not body:
        raise ValueError("header found but no data rows below it")

    cols = map_columns(headers, body[:4])
    if cols.get("text") is None:
        raise ValueError(f"no requirement text column identified in {headers}")

    def cell(row, field):
        i = cols.get(field)
        return row[i] if (i is not None and i < len(row)) else ""

    reqs, derived = [], 0
    for n, row in enumerate(body, 1):
        text = cell(row, "text")
        if len(text) < 10 or looks_like_heading(text, row, cols):
            continue          # section headings and spacer rows
        cid = cell(row, "id")
        if cid:
            rid, origin = cid, "preserved"
        else:
            derived += 1
            rid, origin = f"DER-{derived:03d}", "derived"
        reqs.append({"id": rid, "id_origin": origin, "text": text,
                     "type": cell(row, "type") or "unspecified",
                     "status": cell(row, "status"),
                     "rationale": cell(row, "rationale"),
                     "parent": cell(row, "parent"),
                     "source_row": h + 1 + n})

    return {"requirements": reqs, "sheet": sheet.title, "header_row": h + 1,
            "column_mapping": cols, "headers": headers,
            "preserved_ids": sum(1 for r in reqs if r["id_origin"] == "preserved"),
            "derived_ids": derived}
